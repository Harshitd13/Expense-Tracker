# ============================================================
# Expense Tracker
# A simple tool to track where your money goes
# ============================================================

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
import os
import io
from datetime import datetime, date
import warnings

warnings.filterwarnings("ignore")

# ── Page config ─────────────────────────────────────────────
st.set_page_config(
    page_title="Expense Tracker",
    page_icon="💸",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Constants ────────────────────────────────────────────────
DATA_FILE   = "expenses.csv"
REPORT_DIR  = "reports"
REPORT_FILE = os.path.join(REPORT_DIR, "Expense_Report.xlsx")

CATEGORIES = [
    "Food",
    "Transport",
    "Shopping",
    "Bills",
    "Entertainment",
    "Education",
    "Health",
]

CAT_COLORS = {
    "Food":          "#D87C5A",
    "Transport":     "#7B9EB0",
    "Shopping":      "#C99BB8",
    "Bills":         "#8AAE92",
    "Entertainment": "#E0B66C",
    "Education":     "#9B9BC4",
    "Health":        "#A8C8A6",
}

PAGES = [
    ("Overview",    "🏠"),
    ("Add Expense", "➕"),
    ("Upload CSV",  "📂"),
    ("Clean Data",  "🧹"),
    ("Charts",      "📊"),
    ("Budget",      "🎯"),
    ("Export",      "📥"),
    ("All Records", "📋"),
]


# ── CSS ──────────────────────────────────────────────────────
def apply_css():
    st.markdown(
        """
        <style>

        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

        html, body, [class*="css"] {
            font-family: 'Inter', sans-serif;
            color: #3D2F25;
        }

        .main, .stApp {
            background-color: #FBF7F1;
        }

        .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
        }

        section[data-testid="stSidebar"] {
            background-color: #2E2520 !important;
            border-right: 1px solid #1F1814 !important;
        }

        section[data-testid="stSidebar"] > div:first-child {
            padding-top: 0 !important;
        }

        section[data-testid="stSidebar"] .stButton > button {
            width: 100% !important;
            background-color: transparent !important;
            color: #D5CABF !important;
            border: none !important;
            border-left: 3px solid transparent !important;
            border-radius: 0 8px 8px 0 !important;
            padding: 11px 16px !important;
            text-align: left !important;
            font-size: 14px !important;
            font-weight: 500 !important;
            justify-content: flex-start !important;
            box-shadow: none !important;
            transition: all 0.15s ease !important;
            margin: 2px 0 !important;
            min-height: 42px !important;
        }

        section[data-testid="stSidebar"] .stButton > button:hover {
            background-color: #3D2F25 !important;
            color: #FFFFFF !important;
            border-left: 3px solid #6B5446 !important;
            transform: none !important;
        }

        section[data-testid="stSidebar"] .stButton > button:focus {
            background-color: #3D2F25 !important;
            color: #FFFFFF !important;
            box-shadow: none !important;
            outline: none !important;
        }

        section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
            background-color: #C56E47 !important;
            color: #FFFFFF !important;
            border-left: 3px solid #8B3F1F !important;
            font-weight: 600 !important;
        }

        section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
            background-color: #A85839 !important;
            color: #FFFFFF !important;
            border-left: 3px solid #8B3F1F !important;
        }

        section[data-testid="stSidebar"] hr {
            border-top: 1px solid #3D2F25 !important;
            margin: 14px 16px !important;
        }

        .kpi-box {
            background: #FFFFFF;
            border: 1px solid #EEE5D7;
            border-radius: 14px;
            padding: 22px 20px;
            text-align: left;
            box-shadow: 0 1px 6px rgba(61, 47, 37, 0.06);
            transition: transform 0.15s ease, box-shadow 0.15s ease;
        }

        .kpi-box:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 14px rgba(61, 47, 37, 0.10);
        }

        .kpi-box .top-label {
            font-size: 11px;
            color: #8B7766;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin: 0 0 10px 0;
            font-weight: 600;
        }

        .kpi-box .amount {
            font-size: 26px;
            font-weight: 700;
            color: #3D2F25;
            margin: 0;
            letter-spacing: -0.5px;
        }

        .kpi-box .sub-label {
            font-size: 12px;
            color: #A89683;
            margin: 6px 0 0 0;
        }

        .kpi-box.accent-rust   { border-left: 4px solid #C56E47; }
        .kpi-box.accent-teal   { border-left: 4px solid #7B9EB0; }
        .kpi-box.accent-sage   { border-left: 4px solid #8AAE92; }
        .kpi-box.accent-honey  { border-left: 4px solid #E0B66C; }

        .section-title {
            font-size: 16px;
            font-weight: 600;
            color: #3D2F25;
            padding-bottom: 8px;
            border-bottom: 1.5px solid #EEE5D7;
            margin: 32px 0 18px 0;
        }

        .page-title {
            font-size: 26px;
            font-weight: 700;
            color: #3D2F25;
            margin: 0 0 4px 0;
            letter-spacing: -0.3px;
        }

        .page-sub {
            font-size: 14px;
            color: #8B7766;
            margin: 0 0 24px 0;
        }

        .alert-green {
            background: #F0F5EC;
            border-left: 4px solid #8AAE92;
            border-radius: 8px;
            padding: 16px 20px;
            color: #3D4F35;
            font-size: 14px;
            line-height: 1.6;
        }

        .alert-red {
            background: #FAEEE8;
            border-left: 4px solid #C56E47;
            border-radius: 8px;
            padding: 16px 20px;
            color: #6B3220;
            font-size: 14px;
            line-height: 1.6;
        }

        .alert-yellow {
            background: #FBF1DC;
            border-left: 4px solid #D49B3D;
            border-radius: 8px;
            padding: 16px 20px;
            color: #7A5217;
            font-size: 14px;
            line-height: 1.6;
        }

        .main .stTextInput label,
        .main .stNumberInput label,
        .main .stSelectbox label,
        .main .stMultiSelect label,
        .main .stDateInput label,
        .main .stRadio > label,
        .main .stFileUploader label,
        .main .stTextArea label {
            color: #3D2F25 !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            margin-bottom: 6px !important;
        }

        .main .stTextInput input,
        .main .stNumberInput input,
        .main .stDateInput input,
        .main .stTextArea textarea {
            background-color: #FFFFFF !important;
            color: #3D2F25 !important;
            border: 1.5px solid #E0D5C3 !important;
            border-radius: 8px !important;
            padding: 10px 14px !important;
            font-size: 14px !important;
        }

        .main .stTextInput input:focus,
        .main .stNumberInput input:focus,
        .main .stDateInput input:focus,
        .main .stTextArea textarea:focus {
            border-color: #C56E47 !important;
            box-shadow: 0 0 0 2px rgba(197, 110, 71, 0.15) !important;
            outline: none !important;
        }

        .main .stTextInput input::placeholder,
        .main .stNumberInput input::placeholder,
        .main .stTextArea textarea::placeholder {
            color: #B8A695 !important;
            opacity: 1 !important;
        }

        .main div[data-baseweb="select"] > div {
            background-color: #FFFFFF !important;
            border: 1.5px solid #E0D5C3 !important;
            border-radius: 8px !important;
            color: #3D2F25 !important;
            min-height: 42px !important;
        }

        .main div[data-baseweb="select"] > div:hover {
            border-color: #C56E47 !important;
        }

        .main div[data-baseweb="select"] span {
            color: #3D2F25 !important;
            font-size: 14px !important;
        }

        div[data-baseweb="popover"] {
            background-color: #FFFFFF !important;
        }

        div[data-baseweb="popover"] li {
            background-color: #FFFFFF !important;
            color: #3D2F25 !important;
        }

        div[data-baseweb="popover"] li:hover {
            background-color: #FBF1E8 !important;
        }

        div[data-baseweb="tag"] {
            background-color: #C56E47 !important;
            color: #FFFFFF !important;
            border-radius: 6px !important;
            font-weight: 500 !important;
        }

        div[data-baseweb="tag"] span {
            color: #FFFFFF !important;
        }

        .main .stNumberInput button {
            background-color: #FBF1E8 !important;
            color: #3D2F25 !important;
            border: 1.5px solid #E0D5C3 !important;
            border-radius: 6px !important;
        }

        .main .stNumberInput button:hover {
            background-color: #F5E4D2 !important;
            color: #C56E47 !important;
        }

        .main div[data-baseweb="input"] {
            background-color: #FFFFFF !important;
        }

        .main div[data-baseweb="input"] input {
            color: #3D2F25 !important;
        }

        .main .stRadio div[role="radiogroup"] label p {
            color: #3D2F25 !important;
            font-size: 14px !important;
        }

        .main .stRadio div[role="radiogroup"] label > div:first-child {
            background-color: #FFFFFF !important;
        }

        section[data-testid="stFileUploaderDropzone"] {
            background-color: #FFFFFF !important;
            border: 2px dashed #D4C2AA !important;
            border-radius: 10px !important;
            padding: 24px !important;
        }

        section[data-testid="stFileUploaderDropzone"] * {
            color: #3D2F25 !important;
        }

        section[data-testid="stFileUploaderDropzone"]:hover {
            border-color: #C56E47 !important;
            background-color: #FFFCF8 !important;
        }

        .main .stButton > button,
        .main .stDownloadButton > button,
        .main .stFormSubmitButton > button {
            background-color: #C56E47 !important;
            color: #FFFFFF !important;
            border: none !important;
            border-radius: 8px !important;
            padding: 11px 24px !important;
            font-size: 14px !important;
            font-weight: 600 !important;
            transition: background 0.2s, transform 0.1s !important;
            box-shadow: 0 2px 4px rgba(197, 110, 71, 0.2) !important;
        }

        .main .stButton > button:hover,
        .main .stDownloadButton > button:hover,
        .main .stFormSubmitButton > button:hover {
            background-color: #A85839 !important;
            color: #FFFFFF !important;
            transform: translateY(-1px);
        }

        .stDataFrame, [data-testid="stDataFrame"] {
            border: 1px solid #EEE5D7 !important;
            border-radius: 10px !important;
            overflow: hidden !important;
            background-color: #FFFFFF !important;
        }

        .stDataFrame thead tr th,
        [data-testid="stDataFrame"] thead tr th {
            background-color: #FBF1E8 !important;
            color: #3D2F25 !important;
            font-weight: 600 !important;
            font-size: 13px !important;
            border-bottom: 1.5px solid #E0D5C3 !important;
        }

        .stDataFrame tbody tr td,
        [data-testid="stDataFrame"] tbody tr td {
            background-color: #FFFFFF !important;
            color: #3D2F25 !important;
            font-size: 13px !important;
            border-bottom: 1px solid #F5EDE0 !important;
        }

        .stDataFrame tbody tr:hover td,
        [data-testid="stDataFrame"] tbody tr:hover td {
            background-color: #FBF1E8 !important;
        }

        [data-testid="stMetricValue"] {
            color: #3D2F25 !important;
            font-weight: 700 !important;
        }

        [data-testid="stMetricLabel"] {
            color: #8B7766 !important;
            font-weight: 600 !important;
            font-size: 13px !important;
        }

        [data-testid="stMetricDelta"] {
            color: #8AAE92 !important;
        }

        .streamlit-expanderHeader {
            background-color: #FFFFFF !important;
            border: 1px solid #EEE5D7 !important;
            border-radius: 8px !important;
            color: #3D2F25 !important;
            font-weight: 600 !important;
        }

        .streamlit-expanderContent {
            background-color: #FFFCF8 !important;
            border: 1px solid #EEE5D7 !important;
            border-top: none !important;
            border-radius: 0 0 8px 8px !important;
        }

        div[data-baseweb="notification"] {
            border-radius: 8px !important;
        }

        .stAlert {
            border-radius: 8px !important;
            border-left-width: 4px !important;
        }

        .stTabs [data-baseweb="tab"] {
            color: #8B7766 !important;
            font-weight: 600 !important;
        }

        .stTabs [aria-selected="true"] {
            color: #C56E47 !important;
        }

        .stProgress > div > div {
            background-color: #C56E47 !important;
        }

        .stProgress > div {
            background-color: #EEE5D7 !important;
        }

        .main hr {
            border-color: #EEE5D7 !important;
        }

        .stCaption, [data-testid="stCaptionContainer"] {
            color: #8B7766 !important;
        }

        .footer {
            text-align: center;
            color: #B8A695;
            font-size: 12px;
            padding: 30px 0 10px 0;
            border-top: 1px solid #EEE5D7;
            margin-top: 50px;
        }

        #MainMenu { visibility: hidden; }
        footer    { visibility: hidden; }
        header    { visibility: hidden; }

        ::-webkit-scrollbar {
            width: 10px;
            height: 10px;
        }
        ::-webkit-scrollbar-track {
            background: #FBF7F1;
        }
        ::-webkit-scrollbar-thumb {
            background: #D4C2AA;
            border-radius: 5px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background: #C56E47;
        }
    /* ── Streamlit native alert boxes (info / success / warning / error) ── */
    .main div[data-testid="stAlert"] {
    border-radius: 8px !important;
    border-left-width: 4px !important;
    padding: 14px 18px !important;
    }

    .main div[data-testid="stAlert"] p,
    .main div[data-testid="stAlert"] span,
    .main div[data-testid="stAlert"] div {
        color: #3D2F25 !important;
        font-size: 14px !important;
        line-height: 1.6 !important;
    }

    /* Info (blue by default) → warm cream */
    .main div[data-testid="stAlert"][kind="info"],
    .main div[data-baseweb="notification"][kind="info"] {
        background-color: #FBF1E8 !important;
        border-left: 4px solid #C56E47 !important;
    }

    /* Success → soft sage green */
    .main div[data-testid="stAlert"][kind="success"] {
        background-color: #F0F5EC !important;
        border-left: 4px solid #8AAE92 !important;
    }
    .main div[data-testid="stAlert"][kind="success"] p,
    .main div[data-testid="stAlert"][kind="success"] span {
        color: #3D4F35 !important;
    }

    /* Warning → warm honey */
    .main div[data-testid="stAlert"][kind="warning"] {
        background-color: #FBF1DC !important;
        border-left: 4px solid #D49B3D !important;
    }
    .main div[data-testid="stAlert"][kind="warning"] p,
    .main div[data-testid="stAlert"][kind="warning"] span {
        color: #7A5217 !important;
    }

    /* Error → warm rust */
    .main div[data-testid="stAlert"][kind="error"] {
        background-color: #FAEEE8 !important;
        border-left: 4px solid #C56E47 !important;
    }
    .main div[data-testid="stAlert"][kind="error"] p,
    .main div[data-testid="stAlert"][kind="error"] span {
        color: #6B3220 !important;
    }

    /* Alert icons */
    .main div[data-testid="stAlert"] svg {
        fill: #C56E47 !important;
    }
    .main div[data-testid="stAlert"][kind="success"] svg {
        fill: #8AAE92 !important;
    }
    .main div[data-testid="stAlert"][kind="warning"] svg {
        fill: #D49B3D !important;
    }

        </style>
        """,
        unsafe_allow_html=True,
    )


def load_data() -> pd.DataFrame:
    if os.path.exists(DATA_FILE):
        return pd.read_csv(DATA_FILE)
    df = pd.DataFrame(columns=["Date", "Category", "Description", "Amount"])
    df.to_csv(DATA_FILE, index=False)
    return df


def save_data(df: pd.DataFrame):
    df.to_csv(DATA_FILE, index=False)


def clean_data(df: pd.DataFrame):
    summary = {
        "original_rows":         len(df),
        "empty_rows_removed":    0,
        "duplicates_removed":    0,
        "bad_amounts_removed":   0,
        "bad_dates_removed":     0,
        "final_rows":            0,
    }

    before = len(df)
    df = df.dropna(how="all")
    summary["empty_rows_removed"] = before - len(df)

    df["Description"] = df["Description"].fillna("—")
    df = df.dropna(subset=["Amount", "Category", "Date"])

    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    before = len(df)
    df = df.dropna(subset=["Amount"])
    df = df[df["Amount"] > 0]
    summary["bad_amounts_removed"] = before - len(df)

    before = len(df)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    summary["bad_dates_removed"] = before - len(df)
    df["Date"] = df["Date"].dt.strftime("%Y-%m-%d")

    df["Category"]    = df["Category"].astype(str).str.strip().str.title()
    df["Description"] = df["Description"].astype(str).str.strip()

    before = len(df)
    df = df.drop_duplicates()
    summary["duplicates_removed"] = before - len(df)

    df = df.reset_index(drop=True)
    summary["final_rows"] = len(df)
    return df, summary


def compute_kpis(df: pd.DataFrame) -> dict:
    df = df.copy()
    df["Date"]  = pd.to_datetime(df["Date"])
    df["Month"] = df["Date"].dt.to_period("M").astype(str)

    total          = df["Amount"].sum()
    cat_totals     = df.groupby("Category")["Amount"].sum()
    mon_totals     = df.groupby("Month")["Amount"].sum()
    top_cat        = cat_totals.idxmax() if not cat_totals.empty else "—"
    avg            = df["Amount"].mean() if not df.empty else 0
    cur_mon        = datetime.now().strftime("%Y-%m")
    cur_mon_total  = mon_totals.get(cur_mon, 0)

    if cur_mon_total == 0 and not mon_totals.empty:
        latest_mon = mon_totals.index.max()
        cur_mon_total = mon_totals[latest_mon]
        cur_mon = latest_mon

    return {
        "total":            total,
        "this_month":       cur_mon_total,
        "this_month_label": cur_mon,
        "top_category":     top_cat,
        "average":          avg,
        "cat_totals":       cat_totals,
        "mon_totals":       mon_totals,
    }


def chart_defaults():
    matplotlib.rcParams.update({
        "font.family":        "sans-serif",
        "axes.spines.top":    False,
        "axes.spines.right":  False,
        "axes.grid":          True,
        "grid.color":         "#F0E6D5",
        "grid.linewidth":     0.8,
        "axes.facecolor":     "#FFFCF8",
        "figure.facecolor":   "#FFFCF8",
        "text.color":         "#3D2F25",
        "axes.labelcolor":    "#5A4737",
        "xtick.color":        "#8B7766",
        "ytick.color":        "#8B7766",
        "axes.titleweight":   "600",
        "axes.titlesize":     14,
        "axes.titlepad":      14,
        "axes.titlecolor":    "#3D2F25",
    })


def plot_pie(cat_totals: pd.Series):
    chart_defaults()
    colors = [CAT_COLORS.get(c, "#CCCCCC") for c in cat_totals.index]

    fig, ax = plt.subplots(figsize=(7, 5))
    wedges, _, autotexts = ax.pie(
        cat_totals.values,
        labels=None,
        autopct="%1.0f%%",
        colors=colors,
        startangle=120,
        pctdistance=0.78,
        wedgeprops=dict(width=0.55, edgecolor="#FFFCF8", linewidth=2.5),
    )
    for at in autotexts:
        at.set_fontsize(10)
        at.set_color("#FFFFFF")
        at.set_fontweight("600")

    ax.legend(
        wedges,
        [f"{c}   ₹{v:,.0f}" for c, v in zip(cat_totals.index, cat_totals.values)],
        loc="center left",
        bbox_to_anchor=(1.0, 0.5),
        frameon=False,
        fontsize=10,
        labelcolor="#3D2F25",
    )
    ax.set_title("Where the money goes")
    plt.tight_layout()
    return fig


def plot_bar(cat_totals: pd.Series):
    chart_defaults()
    sorted_totals = cat_totals.sort_values(ascending=True)
    colors = [CAT_COLORS.get(c, "#CCCCCC") for c in sorted_totals.index]

    fig, ax = plt.subplots(figsize=(9, 5))
    bars = ax.barh(
        sorted_totals.index,
        sorted_totals.values,
        color=colors,
        height=0.6,
        edgecolor="none",
    )

    for bar, val in zip(bars, sorted_totals.values):
        ax.text(
            bar.get_width() + sorted_totals.max() * 0.012,
            bar.get_y() + bar.get_height() / 2,
            f"₹{val:,.0f}",
            va="center",
            fontsize=10,
            color="#5A4737",
        )

    ax.set_xlabel("Amount spent (₹)", fontsize=11)
    ax.set_title("Spending by category")
    ax.spines["left"].set_visible(False)
    ax.tick_params(axis="y", left=False)
    ax.grid(axis="x", alpha=0.6)
    ax.grid(axis="y", visible=False)
    plt.tight_layout()
    return fig


def plot_trend(mon_totals: pd.Series):
    chart_defaults()
    months  = mon_totals.index.tolist()
    amounts = mon_totals.values.tolist()

    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(
        months, amounts,
        color="#C56E47",
        linewidth=2.5,
        marker="o",
        markersize=8,
        markerfacecolor="#FFFFFF",
        markeredgecolor="#C56E47",
        markeredgewidth=2.5,
        solid_capstyle="round",
    )
    ax.fill_between(months, amounts, alpha=0.10, color="#C56E47")

    for month, amount in zip(months, amounts):
        ax.annotate(
            f"₹{amount:,.0f}",
            (month, amount),
            xytext=(0, 12),
            textcoords="offset points",
            ha="center",
            fontsize=9,
            color="#5A4737",
        )

    ax.set_xlabel("Month", fontsize=11)
    ax.set_ylabel("Amount (₹)", fontsize=11)
    ax.set_title("Spending over time")
    plt.xticks(rotation=25, ha="right")
    plt.tight_layout()
    return fig


def plot_heatmap(df: pd.DataFrame):
    chart_defaults()
    df = df.copy()
    df["Date"]  = pd.to_datetime(df["Date"])
    df["Month"] = df["Date"].dt.to_period("M").astype(str)
    pivot = df.pivot_table(
        index="Month", columns="Category",
        values="Amount", aggfunc="sum", fill_value=0,
    )
    if pivot.empty:
        return None

    fig, ax = plt.subplots(figsize=(11, max(3.5, len(pivot) * 0.75)))

    from matplotlib.colors import LinearSegmentedColormap
    warm_cmap = LinearSegmentedColormap.from_list(
        "warm", ["#FBF1E8", "#E8B896", "#C56E47", "#8B3F1F"]
    )

    im = ax.imshow(pivot.values, cmap=warm_cmap, aspect="auto")

    ax.set_xticks(range(len(pivot.columns)))
    ax.set_xticklabels(pivot.columns, rotation=25, ha="right", fontsize=10)
    ax.set_yticks(range(len(pivot.index)))
    ax.set_yticklabels(pivot.index, fontsize=10)
    ax.spines[:].set_visible(False)
    ax.tick_params(length=0)

    max_val = pivot.values.max()
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            val = pivot.values[i][j]
            if val > 0:
                ax.text(
                    j, i,
                    f"₹{val:,.0f}",
                    ha="center", va="center",
                    fontsize=8,
                    color="#FFFFFF" if val > max_val * 0.55 else "#5A4737",
                )

    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label("Amount (₹)", color="#5A4737")
    cbar.ax.tick_params(colors="#8B7766")

    ax.set_title("Month-by-category breakdown")
    plt.tight_layout()
    return fig


def export_to_excel(df: pd.DataFrame) -> str:
    os.makedirs(REPORT_DIR, exist_ok=True)

    df_copy = df.copy()
    df_copy["Date"]  = pd.to_datetime(df_copy["Date"])
    df_copy["Month"] = df_copy["Date"].dt.to_period("M").astype(str)

    cat_sum = (
        df_copy.groupby("Category")["Amount"]
        .agg(Total="sum", Average="mean", Transactions="count")
        .reset_index()
        .sort_values("Total", ascending=False)
    )
    mon_sum = (
        df_copy.groupby("Month")["Amount"]
        .agg(Total="sum", Average="mean", Transactions="count")
        .reset_index()
    )

    with pd.ExcelWriter(REPORT_FILE, engine="openpyxl") as writer:
        df_copy[["Date", "Category", "Description", "Amount"]].to_excel(
            writer, sheet_name="All Expenses", index=False
        )
        cat_sum.to_excel(writer, sheet_name="By Category", index=False)
        mon_sum.to_excel(writer, sheet_name="By Month",    index=False)

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        H_FILL   = PatternFill(start_color="C56E47", end_color="C56E47", fill_type="solid")
        H_FONT   = Font(color="FFFFFF", bold=True, size=11)
        ALT_FILL = PatternFill(start_color="FBF1E8", end_color="FBF1E8", fill_type="solid")
        CENTER   = Alignment(horizontal="center", vertical="center")
        THIN     = Side(style="thin", color="E0D5C3")
        BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        for sname in writer.sheets:
            ws = writer.sheets[sname]

            for cell in ws[1]:
                cell.fill      = H_FILL
                cell.font      = H_FONT
                cell.alignment = CENTER
                cell.border    = BORDER

            for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                for cell in row:
                    cell.border    = BORDER
                    cell.alignment = CENTER
                    if ridx % 2 == 0:
                        cell.fill = ALT_FILL

            for cidx, col in enumerate(ws.columns, start=1):
                width = max(
                    (len(str(c.value)) for c in col if c.value), default=10
                )
                ws.column_dimensions[get_column_letter(cidx)].width = min(
                    width + 5, 42
                )

    return REPORT_FILE


def render_sidebar() -> str:
    if "current_page" not in st.session_state:
        st.session_state.current_page = "Overview"

    with st.sidebar:

        st.markdown(
            "<div style='padding:24px 16px 12px 16px;text-align:left;'>"
            "<p style='font-size:32px;margin:0;line-height:1;'>💸</p>"
            "<p style='color:#FFFFFF;font-size:19px;font-weight:700;margin:12px 0 4px 0;letter-spacing:-0.3px;'>Expense Tracker</p>"
            "<p style='color:#A08B7A;font-size:12px;margin:0;font-weight:400;'>Track where your money goes</p>"
            "</div>",
            unsafe_allow_html=True,
        )

        st.divider()

        st.markdown(
            "<p style='color:#7A6859;font-size:10px;text-transform:uppercase;letter-spacing:2px;margin:0 16px 8px 16px;font-weight:700;'>Menu</p>",
            unsafe_allow_html=True,
        )

        for page_name, icon in PAGES:
            is_active = (st.session_state.current_page == page_name)
            button_type = "primary" if is_active else "secondary"

            if st.button(
                f"{icon}   {page_name}",
                key=f"nav_{page_name}",
                type=button_type,
                use_container_width=True,
            ):
                st.session_state.current_page = page_name
                st.rerun()

        st.divider()

        df = load_data()
        if not df.empty:
            df_c, _ = clean_data(df.copy())
            if not df_c.empty:
                total = df_c["Amount"].sum()
                count = len(df_c)

                st.markdown(
                    "<p style='color:#7A6859;font-size:10px;text-transform:uppercase;letter-spacing:2px;margin:0 0 12px 8px;font-weight:700;'>Your Numbers</p>",
                    unsafe_allow_html=True,
                )

                st.markdown(
                    f"<div style='background:#3A2F28;border-radius:12px;padding:18px 16px;border-left:3px solid #C56E47;margin:0 4px;'>"
                    f"<p style='color:#8B7766;font-size:10px;text-transform:uppercase;letter-spacing:1.2px;margin:0 0 8px 0;font-weight:600;'>Total Spent</p>"
                    f"<p style='color:#FFFFFF;font-size:24px;font-weight:700;margin:0;letter-spacing:-0.5px;line-height:1.2;'>₹{total:,.0f}</p>"
                    f"<p style='color:#E8A47C;font-size:12px;margin:8px 0 0 0;font-weight:500;'>{count} transactions recorded</p>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

        st.markdown(
            "<div style='margin-top:40px;padding:16px 0 12px 0;text-align:center;border-top:1px solid #3D2F25;'>"
            "<p style='color:#54453D;font-size:11px;margin:0;font-weight:500;'>Built with Python · Streamlit</p>"
            "</div>",
            unsafe_allow_html=True,
        )

    return st.session_state.current_page


def kpi_card(label, value, sub, accent_class):
    return (
        f"<div class='kpi-box {accent_class}'>"
        f"<p class='top-label'>{label}</p>"
        f"<p class='amount'>{value}</p>"
        f"<p class='sub-label'>{sub}</p>"
        f"</div>"
    )


def page_overview():
    st.markdown(
        "<p class='page-title'>Overview</p>"
        "<p class='page-sub'>A snapshot of your spending so far.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("No expenses yet. Head to **Add Expense** in the sidebar to begin.")
        return

    df_c, _ = clean_data(df.copy())
    if df_c.empty:
        st.warning("Data loaded but nothing valid after cleaning. Check your CSV.")
        return

    kpis = compute_kpis(df_c.copy())

    st.markdown("<p class='section-title'>At a glance</p>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)

    c1.markdown(kpi_card("Total spent", f"₹{kpis['total']:,.0f}",
                         "all time", "accent-rust"), unsafe_allow_html=True)
    c2.markdown(kpi_card("This month", f"₹{kpis['this_month']:,.0f}",
                         kpis["this_month_label"], "accent-teal"), unsafe_allow_html=True)
    c3.markdown(kpi_card("Top category", kpis["top_category"],
                         "biggest spender", "accent-sage"), unsafe_allow_html=True)
    c4.markdown(kpi_card("Avg per entry", f"₹{kpis['average']:,.0f}",
                         "average transaction", "accent-honey"), unsafe_allow_html=True)

    st.markdown("<p class='section-title'>Where it's going</p>", unsafe_allow_html=True)
    col_l, col_r = st.columns(2)
    with col_l:
        fig = plot_pie(kpis["cat_totals"])
        st.pyplot(fig); plt.close(fig)
    with col_r:
        fig = plot_bar(kpis["cat_totals"])
        st.pyplot(fig); plt.close(fig)

    st.markdown("<p class='section-title'>Spending over time</p>", unsafe_allow_html=True)
    fig = plot_trend(kpis["mon_totals"])
    st.pyplot(fig); plt.close(fig)

    st.markdown("<p class='section-title'>Recent entries</p>", unsafe_allow_html=True)
    recent = df_c.copy()
    recent["Date"] = pd.to_datetime(recent["Date"])
    recent = recent.sort_values("Date", ascending=False).head(8)
    recent["Date"]   = recent["Date"].dt.strftime("%d %b %Y")
    recent["Amount"] = recent["Amount"].apply(lambda x: f"₹{x:,.2f}")
    st.dataframe(recent, use_container_width=True, hide_index=True)


def page_add_expense():
    st.markdown(
        "<p class='page-title'>Add an Expense</p>"
        "<p class='page-sub'>Quick form — takes about 10 seconds.</p>",
        unsafe_allow_html=True,
    )

    st.info(
        "💡 Tip: be specific in the description. "
        "\"Coffee with Aman\" is more useful than just \"Coffee\"."
    )

    st.markdown("<br>", unsafe_allow_html=True)

    with st.form("add_form", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            exp_date = st.date_input("Date", value=date.today())
            category = st.selectbox("Category", CATEGORIES)

        with c2:
            amount = st.number_input(
                "Amount (₹)",
                min_value=0.0,
                max_value=999_999.0,
                value=0.0,
                step=10.0,
            )
            description = st.text_input(
                "Description",
                placeholder="What did you spend on?",
            )

        st.markdown("<br>", unsafe_allow_html=True)
        save_btn = st.form_submit_button("Save expense", use_container_width=True)

        if save_btn:
            if not description or not description.strip():
                st.error("Please add a short description.")
            elif amount <= 0:
                st.error("Enter an amount greater than zero.")
            else:
                df = load_data()
                new = pd.DataFrame({
                    "Date":        [exp_date.strftime("%Y-%m-%d")],
                    "Category":    [category],
                    "Description": [description.strip()],
                    "Amount":      [round(float(amount), 2)],
                })
                df = pd.concat([df, new], ignore_index=True)
                save_data(df)
                st.success(
                    f"Saved — ₹{amount:,.2f} for {category} on "
                    f"{exp_date.strftime('%d %b %Y')}"
                )

    st.markdown("<p class='section-title'>Today's entries</p>", unsafe_allow_html=True)
    df = load_data()
    if not df.empty:
        df["Date"] = pd.to_datetime(df["Date"])
        today_df = df[df["Date"].dt.date == date.today()].copy()
        if not today_df.empty:
            total_today = today_df["Amount"].sum()
            today_df["Amount"] = today_df["Amount"].apply(lambda x: f"₹{x:,.2f}")
            today_df["Date"]   = today_df["Date"].dt.strftime("%d %b %Y")
            st.dataframe(today_df, use_container_width=True, hide_index=True)
            st.caption(f"₹{total_today:,.2f} spent today across {len(today_df)} entries.")
        else:
            st.caption("Nothing recorded today yet.")
    else:
        st.caption("No expense data found.")


def page_upload_csv():
    st.markdown(
        "<p class='page-title'>Upload a CSV</p>"
        "<p class='page-sub'>Import expenses from a spreadsheet.</p>",
        unsafe_allow_html=True,
    )

    with st.expander("What format does my CSV need?"):
        st.markdown(
            "Your file needs these **4 columns** (header names matter):\n\n"
            "| Column | Format | Example |\n"
            "|---|---|---|\n"
            "| Date | YYYY-MM-DD | 2024-03-15 |\n"
            "| Category | Text | Food |\n"
            "| Description | Text | Weekly groceries |\n"
            "| Amount | Number | 1250.00 |"
        )

    st.markdown("<br>", unsafe_allow_html=True)
    uploaded = st.file_uploader("Choose CSV file", type=["csv"])

    if uploaded:
        try:
            df_up = pd.read_csv(uploaded)
        except Exception as e:
            st.error(f"Couldn't read that file. Error: {e}")
            return

        st.markdown("<p class='section-title'>Preview</p>", unsafe_allow_html=True)
        st.dataframe(df_up.head(10), use_container_width=True, hide_index=True)

        required = ["Date", "Category", "Description", "Amount"]
        missing  = [c for c in required if c not in df_up.columns]
        if missing:
            st.error(f"Missing columns: {', '.join(missing)}")
            return

        c1, c2, c3 = st.columns(3)
        c1.metric("Rows", len(df_up))
        c2.metric("Valid amounts",
                  int(pd.to_numeric(df_up["Amount"], errors="coerce").notna().sum()))
        c3.metric("Valid dates",
                  int(pd.to_datetime(df_up["Date"], errors="coerce").notna().sum()))

        st.success("File looks good. Choose how to import.")
        st.markdown("<br>", unsafe_allow_html=True)

        mode = st.radio("Import mode",
                        ["Add to existing data", "Replace all existing data"])

        if st.button("Import now", use_container_width=True):
            if "Replace" in mode:
                save_data(df_up)
                st.success(f"Done — replaced everything with {len(df_up)} rows.")
            else:
                existing = load_data()
                merged   = pd.concat([existing, df_up], ignore_index=True)
                save_data(merged)
                st.success(
                    f"Added {len(df_up)} rows. "
                    f"You now have {len(merged)} records total."
                )


def page_clean_data():
    st.markdown(
        "<p class='page-title'>Clean Data</p>"
        "<p class='page-sub'>Remove duplicates, fix bad dates, drop invalid amounts.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("No data to clean yet.")
        return

    st.markdown("<p class='section-title'>Current state</p>", unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total rows", len(df))
    c2.metric("Missing values", int(df.isnull().sum().sum()))
    c3.metric("Duplicate rows", int(df.duplicated().sum()))
    c4.metric("Columns", len(df.columns))

    with st.expander("What does the cleaner do?"):
        st.markdown(
            "1. Drops completely empty rows\n"
            "2. Fills blank descriptions with a dash\n"
            "3. Removes rows missing date, category, or amount\n"
            "4. Converts amounts to numbers, removes anything ≤ 0\n"
            "5. Parses dates, drops unparseable rows\n"
            "6. Trims whitespace and normalises category capitalisation\n"
            "7. Removes exact duplicate rows"
        )

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("Run cleaner", use_container_width=True):
        with st.spinner("Cleaning …"):
            df_clean, summary = clean_data(df.copy())

        st.markdown("<p class='section-title'>Results</p>", unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Started with", summary["original_rows"])
        c2.metric("Empty removed", summary["empty_rows_removed"])
        c3.metric("Duplicates removed", summary["duplicates_removed"])
        c4.metric("Clean rows", summary["final_rows"])

        rows_removed = summary["original_rows"] - summary["final_rows"]
        if rows_removed == 0:
            st.success("Your data is already clean. Nothing was removed.")
        else:
            st.warning(
                f"Removed {rows_removed} problematic row(s). "
                f"{summary['final_rows']} clean rows remain."
            )

        st.markdown("<p class='section-title'>Preview (first 15 rows)</p>",
                    unsafe_allow_html=True)
        st.dataframe(df_clean.head(15), use_container_width=True, hide_index=True)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Save cleaned data", use_container_width=True):
            save_data(df_clean)
            st.success("Saved. Your data file has been updated.")


def page_charts():
    st.markdown(
        "<p class='page-title'>Charts</p>"
        "<p class='page-sub'>Visualise your spending however you like.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("No data to chart. Add some expenses first.")
        return

    df_c, _ = clean_data(df.copy())
    if df_c.empty:
        st.warning("Nothing usable after cleaning.")
        return

    df_c["Date"] = pd.to_datetime(df_c["Date"])

    st.markdown("<p class='section-title'>Filters</p>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        sel_cats = st.multiselect("Categories", CATEGORIES, default=CATEGORIES)
    with c2:
        min_d = df_c["Date"].min().date()
        max_d = df_c["Date"].max().date()
        date_range = st.date_input(
            "Date range",
            value=(min_d, max_d),
            min_value=min_d,
            max_value=max_d,
        )

    filtered = df_c[df_c["Category"].isin(sel_cats)].copy()
    if len(date_range) == 2:
        filtered = filtered[
            (filtered["Date"].dt.date >= date_range[0])
            & (filtered["Date"].dt.date <= date_range[1])
        ]

    if filtered.empty:
        st.warning("No data matches those filters.")
        return

    cat_totals = filtered.groupby("Category")["Amount"].sum()
    filtered["Month"] = filtered["Date"].dt.to_period("M").astype(str)
    mon_totals = filtered.groupby("Month")["Amount"].sum()

    st.markdown("<p class='section-title'>Category split</p>", unsafe_allow_html=True)
    col_l, col_r = st.columns([3, 2])
    with col_l:
        fig = plot_pie(cat_totals)
        st.pyplot(fig); plt.close(fig)
    with col_r:
        st.markdown(
            "<p style='font-size:13px;font-weight:600;color:#3D2F25;margin:0 0 8px 0;'>Breakdown</p>",
            unsafe_allow_html=True,
        )
        total_all = cat_totals.sum()
        for cat, amt in cat_totals.sort_values(ascending=False).items():
            pct   = (amt / total_all) * 100
            color = CAT_COLORS.get(cat, "#CCC")
            st.markdown(
                f"<div style='display:flex;justify-content:space-between;"
                f"align-items:center;padding:9px 14px;"
                f"background:#FFFFFF;"
                f"border-radius:0 8px 8px 0;margin:5px 0;"
                f"border:1px solid #EEE5D7;border-left:3px solid {color};'>"
                f"<span style='font-size:13px;color:#3D2F25;font-weight:500;'>{cat}</span>"
                f"<span style='font-size:13px;color:#5A4737;'>"
                f"₹{amt:,.0f} &nbsp;<span style='color:#A89683;font-size:11px;'>"
                f"({pct:.0f}%)</span></span></div>",
                unsafe_allow_html=True,
            )

    st.markdown("<p class='section-title'>Category totals</p>", unsafe_allow_html=True)
    fig = plot_bar(cat_totals)
    st.pyplot(fig); plt.close(fig)

    st.markdown("<p class='section-title'>Monthly trend</p>", unsafe_allow_html=True)
    fig = plot_trend(mon_totals)
    st.pyplot(fig); plt.close(fig)

    st.markdown("<p class='section-title'>Heatmap</p>", unsafe_allow_html=True)
    fig = plot_heatmap(filtered)
    if fig:
        st.pyplot(fig); plt.close(fig)


def page_budget():
    st.markdown(
        "<p class='page-title'>Budget</p>"
        "<p class='page-sub'>Set a limit, see how you're tracking.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("No data yet. Add some expenses first.")
        return

    df_c, _ = clean_data(df.copy())
    df_c["Date"]  = pd.to_datetime(df_c["Date"])
    df_c["Month"] = df_c["Date"].dt.to_period("M").astype(str)
    all_months    = sorted(df_c["Month"].unique(), reverse=True)

    st.markdown("<p class='section-title'>Settings</p>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        budget = st.number_input(
            "Monthly budget (₹)",
            min_value=500.0,
            max_value=999_999.0,
            value=15_000.0,
            step=500.0,
        )
    with c2:
        sel_month = st.selectbox("Month to check", all_months)

    with st.expander("Set per-category limits (optional)"):
        cat_budgets = {}
        cols = st.columns(4)
        for i, cat in enumerate(CATEGORIES):
            with cols[i % 4]:
                cat_budgets[cat] = st.number_input(
                    cat,
                    min_value=0.0,
                    max_value=50_000.0,
                    value=0.0,
                    step=100.0,
                    key=f"cb_{cat}",
                )

    st.markdown("<br>", unsafe_allow_html=True)

    month_df    = df_c[df_c["Month"] == sel_month]
    month_spent = month_df["Amount"].sum()
    remaining   = budget - month_spent
    pct_used    = (month_spent / budget * 100) if budget else 0

    st.markdown(
        f"<p class='section-title'>{sel_month} results</p>",
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    c1.metric("Budget", f"₹{budget:,.0f}")
    c2.metric("Spent",  f"₹{month_spent:,.0f}")
    c3.metric(
        "Left" if remaining >= 0 else "Over by",
        f"₹{abs(remaining):,.0f}",
    )

    st.markdown("<br>", unsafe_allow_html=True)

    if month_spent > budget:
        biggest = (
            df_c[df_c['Month'] == sel_month]
            .groupby('Category')['Amount']
            .sum()
            .idxmax()
        )
        st.markdown(
            f"<div class='alert-red'>"
            f"<strong>Over budget by ₹{abs(remaining):,.0f}</strong><br>"
            f"You spent {pct_used:.0f}% of your ₹{budget:,.0f} budget this month. "
            f"Try cutting back on <strong>{biggest}</strong>."
            f"</div>",
            unsafe_allow_html=True,
        )
    elif pct_used >= 80:
        st.markdown(
            f"<div class='alert-yellow'>"
            f"<strong>Getting close — {pct_used:.0f}% used</strong><br>"
            f"You have ₹{remaining:,.0f} left for the rest of {sel_month}. "
            f"Keep an eye on your spending."
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"<div class='alert-green'>"
            f"<strong>Within budget — {pct_used:.0f}% used</strong><br>"
            f"₹{remaining:,.0f} still available for {sel_month}. Looking good."
            f"</div>",
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    st.progress(min(pct_used / 100, 1.0))
    st.caption(f"₹{month_spent:,.0f} of ₹{budget:,.0f} ({pct_used:.1f}%)")

    st.markdown("<p class='section-title'>Category breakdown</p>",
                unsafe_allow_html=True)
    if not month_df.empty:
        cat_breakdown = (
            month_df.groupby("Category")["Amount"]
            .sum()
            .sort_values(ascending=False)
            .reset_index()
        )
        cat_breakdown.columns = ["Category", "Spent"]
        cat_breakdown["% of total"] = (
            cat_breakdown["Spent"] / month_spent * 100
        ).round(1)
        cat_breakdown["Spent"] = cat_breakdown["Spent"].apply(
            lambda x: f"₹{x:,.2f}"
        )
        cat_breakdown["% of total"] = cat_breakdown["% of total"].apply(
            lambda x: f"{x}%"
        )
        st.dataframe(cat_breakdown, use_container_width=True, hide_index=True)

    active = {c: b for c, b in cat_budgets.items() if b > 0}
    if active:
        st.markdown("<p class='section-title'>Category limits</p>",
                    unsafe_allow_html=True)
        cat_spent_map = month_df.groupby("Category")["Amount"].sum()
        for cat, cat_budget in active.items():
            spent_cat = cat_spent_map.get(cat, 0)
            pct_cat   = (spent_cat / cat_budget * 100) if cat_budget else 0
            c1, c2, c3 = st.columns([2, 2, 4])
            c1.markdown(f"**{cat}**")
            c2.markdown(f"₹{spent_cat:,.0f} / ₹{cat_budget:,.0f}")
            c3.progress(min(pct_cat / 100, 1.0), text=f"{pct_cat:.0f}%")

    st.markdown("<p class='section-title'>All months vs budget</p>",
                unsafe_allow_html=True)
    all_sum = df_c.groupby("Month")["Amount"].sum().reset_index()
    all_sum.columns = ["Month", "Spent"]
    all_sum["Budget"]    = budget
    all_sum["Remaining"] = all_sum["Budget"] - all_sum["Spent"]
    all_sum["Status"]    = all_sum["Remaining"].apply(
        lambda x: "Within budget" if x >= 0 else "Over budget"
    )
    all_sum["Spent"]     = all_sum["Spent"].apply(lambda x: f"₹{x:,.2f}")
    all_sum["Budget"]    = all_sum["Budget"].apply(lambda x: f"₹{x:,.2f}")
    all_sum["Remaining"] = all_sum["Remaining"].apply(lambda x: f"₹{abs(x):,.2f}")
    st.dataframe(all_sum, use_container_width=True, hide_index=True)


def page_export():
    st.markdown(
        "<p class='page-title'>Export</p>"
        "<p class='page-sub'>Download your data as Excel or CSV.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("Nothing to export yet.")
        return

    df_c, _ = clean_data(df.copy())

    st.markdown("<p class='section-title'>Excel report</p>", unsafe_allow_html=True)

    st.markdown(
        "The Excel file includes three sheets:\n\n"
        "- **All Expenses** — every transaction\n"
        "- **By Category** — totals and averages per category\n"
        "- **By Month** — same summary, grouped by month"
    )

    st.markdown("<br>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    c1.metric("Rows to export", len(df_c))
    c2.metric("Categories", df_c["Category"].nunique())
    c3.metric("Total amount", f"₹{df_c['Amount'].sum():,.0f}")

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("Generate Excel file", use_container_width=True):
        with st.spinner("Building report …"):
            path = export_to_excel(df_c)
        st.success("Your file is ready.")
        with open(path, "rb") as f:
            st.download_button(
                label="Download Expense_Report.xlsx",
                data=f.read(),
                file_name="Expense_Report.xlsx",
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

    st.markdown("<p class='section-title'>CSV export</p>", unsafe_allow_html=True)
    buf = io.StringIO()
    df_c.to_csv(buf, index=False)
    st.download_button(
        label="Download as CSV",
        data=buf.getvalue(),
        file_name="my_expenses.csv",
        mime="text/csv",
        use_container_width=True,
    )


def page_all_records():
    st.markdown(
        "<p class='page-title'>All Records</p>"
        "<p class='page-sub'>Browse, search, and filter every entry.</p>",
        unsafe_allow_html=True,
    )

    df = load_data()
    if df.empty:
        st.info("No records yet.")
        return

    df_c, _ = clean_data(df.copy())
    df_c["Date"] = pd.to_datetime(df_c["Date"])

    st.markdown("<p class='section-title'>Filter</p>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        f_cats = st.multiselect("Category", CATEGORIES, default=CATEGORIES)
    with c2:
        f_min = st.number_input("Min amount (₹)", min_value=0.0, value=0.0, step=10.0)
    with c3:
        max_val = float(df_c["Amount"].max()) if not df_c.empty else 10000.0
        f_max = st.number_input("Max amount (₹)", min_value=0.0, value=max_val, step=10.0)

    search = st.text_input("Search description", placeholder="Type anything …")

    filtered = df_c[
        df_c["Category"].isin(f_cats)
        & (df_c["Amount"] >= f_min)
        & (df_c["Amount"] <= f_max)
    ].copy()

    if search.strip():
        filtered = filtered[
            filtered["Description"].str.contains(search.strip(), case=False, na=False)
        ]

    c1, c2 = st.columns(2)
    with c1:
        sort_by  = st.selectbox("Sort by", ["Date", "Amount", "Category"])
    with c2:
        sort_asc = st.radio("Order", ["Newest first", "Oldest first"], horizontal=True)

    filtered = filtered.sort_values(sort_by, ascending=(sort_asc == "Oldest first"))

    st.markdown(
        f"<p class='section-title'>{len(filtered)} record(s) found</p>",
        unsafe_allow_html=True,
    )

    if filtered.empty:
        st.info("No records match your filters. Try adjusting them.")
    else:
        disp = filtered.copy()
        disp["Date"]   = disp["Date"].dt.strftime("%d %b %Y")
        disp["Amount"] = disp["Amount"].apply(lambda x: f"₹{x:,.2f}")
        st.dataframe(disp, use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("⚠️ Delete all data"):
        st.warning(
            "This is permanent. All records will be wiped. "
            "Export first if you need a backup."
        )
        confirm = st.text_input('Type "delete everything" to confirm:')
        if st.button("Yes, delete all records", use_container_width=True):
            if confirm.strip().lower() == "delete everything":
                save_data(
                    pd.DataFrame(
                        columns=["Date", "Category", "Description", "Amount"]
                    )
                )
                st.success("All records deleted.")
                st.rerun()
            else:
                st.error('Please type "delete everything" exactly.')


def main():
    apply_css()

    page = render_sidebar()

    routes = {
        "Overview":    page_overview,
        "Add Expense": page_add_expense,
        "Upload CSV":  page_upload_csv,
        "Clean Data":  page_clean_data,
        "Charts":      page_charts,
        "Budget":      page_budget,
        "Export":      page_export,
        "All Records": page_all_records,
    }

    routes[page]()

    st.markdown(
        "<div class='footer'>"
        "Expense Tracker &nbsp;·&nbsp; "
        "made with Python, Streamlit, and Pandas"
        "</div>",
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()